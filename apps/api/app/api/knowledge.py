from __future__ import annotations

import csv
import hashlib
import io
import zipfile
from datetime import datetime
from pathlib import Path
from uuid import UUID, uuid4
from xml.etree import ElementTree

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel, Field

from app.core.permissions import require_permission
from app.core.tenant_context import get_current_tenant
from app.knowledge import repository as repo
from app.knowledge import pilot_repository as pilot_repo
from app.services.dossier_document_storage import create_document_download_url, upload_private_document
from app.services.knowledge_security import scan_bytes
from app.services.knowledge_ai import ocr_document
from app.knowledge.live_sources import catalog as live_catalog,resolve as resolve_live

router = APIRouter(prefix="/knowledge", tags=["knowledge"])
BUCKET = "knowledge-files"
MIMES = {"application/pdf", "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "text/csv", "text/plain", "image/jpeg", "image/png", "image/webp"}
MAX_FILE_SIZE = 20 * 1024 * 1024


def aid(tenant): return str(tenant.get("user_id") or "system")
def aname(tenant): return str(tenant.get("actor_name") or "Membre de l'agence")


class CreateKnowledge(BaseModel):
    title: str = Field(min_length=2, max_length=240)
    knowledge_type: str = "TEXT"
    category: str = "OTHER"
    content: str = ""
    structured_data: dict = {}
    question_variants: list[str] = []
    tags: list[str] = []
    language: str = "FR"
    audiences: list[str] = ["EMPLOYEES"]
    ai_scope: str = "NONE"
    source_type: str = "MANUAL"
    source_entity_type: str | None = None
    source_entity_id: str | None = None
    effective_at: datetime | None = None
    expires_at: datetime | None = None
    review_due_at: datetime | None = None
    review_interval_days: int | None = Field(default=None, gt=0)
    owner_id: str | None = None
    owner_name: str | None = None
    sensitive: bool = False
    workspace_id: str | None = None


class UpdateKnowledge(BaseModel):
    expected_version: int = Field(ge=1)
    change_reason: str | None = None
    title: str | None = Field(default=None, min_length=2, max_length=240)
    knowledge_type: str | None = None
    category: str | None = None
    content: str | None = None
    structured_data: dict | None = None
    question_variants: list[str] | None = None
    tags: list[str] | None = None
    language: str | None = None
    audiences: list[str] | None = None
    ai_scope: str | None = None
    source_type: str | None = None
    source_entity_type: str | None = None
    source_entity_id: str | None = None
    effective_at: datetime | None = None
    expires_at: datetime | None = None
    review_due_at: datetime | None = None
    review_interval_days: int | None = None
    owner_id: str | None = None
    owner_name: str | None = None
    sensitive: bool | None = None
    workspace_id: str | None = None


class Action(BaseModel):
    reason: str | None = None


class Playground(BaseModel):
    question: str = Field(min_length=2, max_length=2000)
    channel: str = "PLAYGROUND"
    language: str = "FR"
    workspace_id: str | None = None
    limit: int = Field(default=8, ge=1, le=20)
    context: dict = {}


class Settings(BaseModel):
    client_ai_enabled: bool | None = None
    internal_ai_enabled: bool | None = None
    default_language: str | None = None
    response_tone: str | None = None
    escalation_topics: list[str] | None = None
    client_fallback_message: str | None = None
    system_rules: list[str] | None = None
    retention_days: int | None = Field(default=None, ge=30, le=3650)


class ImportFile(BaseModel):
    title: str
    knowledge_type: str = "DOCUMENT"
    category: str = "DOCUMENTS"
    content: str | None = None
    structured_data: dict = {}
    tags: list[str] = []
    language: str = "FR"
    audiences: list[str] = ["EMPLOYEES"]
    owner_name: str | None = None
    review_interval_days: int | None = None
    sensitive: bool = False


class Relation(BaseModel):
    entity_type: str
    entity_id: str
    relation_type: str = "APPLIES_TO"


class ConflictResolution(BaseModel):
    resolution: str = Field(min_length=3)
    status: str = "RESOLVED"


class Feedback(BaseModel):
    knowledge_id: str | None = None
    response_log_id: str | None = None
    rating: str
    comment: str | None = None
class SavedView(BaseModel):name:str=Field(min_length=2,max_length=80);filters:dict={}
class Translation(BaseModel):target_language:str=Field(pattern="^(FR|EN)$")
class Connector(BaseModel):provider:str=Field(pattern="^(GOOGLE_DRIVE|NOTION|SHAREPOINT)$");display_name:str;credentials:dict;configuration:dict={};workspace_id:str|None=None
class SuggestionDecision(BaseModel):
    status: str = Field(pattern="^(ACCEPTED|DISMISSED|COMPLETED)$")
    knowledge_id: str | None = None


class PilotKnowledgePayload(BaseModel):
    subject: str = Field(min_length=2, max_length=240)
    answer: str = Field(min_length=2, max_length=250000)
    kind: str = Field(pattern="^(CLIENT_ANSWER|COMPANY_INFORMATION|INTERNAL_INSTRUCTION)$")
    category: str = Field(default="OTHER", pattern="^(GENERAL|SERVICES|HOURS_AND_ADDRESSES|PAYMENTS|DOSSIERS|SUPPORT|DOCUMENTS|OTHER)$")
    client_visible: bool = True
    language: str = Field(default="FR", pattern="^(FR|EN)$")
    review_due_at: datetime | None = None
    idempotency_key: str | None = Field(default=None, max_length=180)
    source_file_id: UUID | None = None


class PilotKnowledgeUpdate(PilotKnowledgePayload):
    expected_version: int = Field(ge=1)


class PilotKnowledgeAction(BaseModel):
    expected_version: int = Field(ge=1)


@router.get("")
def index(q: str | None = None, status: str | None = None, category: str | None = None, knowledge_type: str | None = None, language: str | None = None, source_type: str | None = None, ai_scope: str | None = None, audience: str | None = None, workspace_id: str | None = None, expired: bool = False, limit: int = 50, offset: int = 0, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.read"))):
    return repo.listing(tenant["org_id"], locals())


@router.get("/stats")
def knowledge_stats(tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.read"))): return repo.stats(tenant["org_id"])


@router.get("/analytics")
def knowledge_analytics(tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.analytics"))): return repo.analytics(tenant["org_id"])


@router.get("/settings")
def get_settings(tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.read"))): return repo.settings(tenant["org_id"])

@router.get("/views")
def views(tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.read"))):return {"items":repo.saved_views(tenant["org_id"],aid(tenant))}
@router.post("/views")
def view_create(body:SavedView,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.read"))):return repo.save_view(tenant["org_id"],aid(tenant),body.name,body.filters)
@router.delete("/views/{view_id}")
def view_delete(view_id:str,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.read"))):return {"deleted":repo.delete_view(tenant["org_id"],aid(tenant),view_id)}

@router.get("/suggestions")
def suggestion_list(tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.read"))):return {"items":repo.suggestions(tenant["org_id"])}
@router.post("/suggestions/generate")
def suggestion_generate(tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.manage"))):return repo.generate_suggestions(tenant["org_id"])
@router.patch("/suggestions/{suggestion_id}")
def suggestion_decide(suggestion_id:str,body:SuggestionDecision,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.manage"))):return repo.update_suggestion(tenant["org_id"],suggestion_id,body.status,body.knowledge_id)

@router.get("/connectors")
def connector_list(tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.connectors"))):return {"items":repo.connectors(tenant["org_id"])}
@router.post("/connectors")
def connector_create(body:Connector,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.connectors"))):return repo.create_connector(tenant["org_id"],aid(tenant),body.model_dump())
@router.post("/connectors/{connector_id}/sync")
def connector_sync(connector_id:str,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.connectors"))):return repo.sync_connector(tenant["org_id"],connector_id)
@router.delete("/connectors/{connector_id}")
def connector_delete(connector_id:str,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.connectors"))):return {"deleted":repo.delete_connector(tenant["org_id"],connector_id)}


@router.patch("/settings")
def patch_settings(body: Settings, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.manage"))): return repo.update_settings(tenant["org_id"], aid(tenant), body.model_dump(exclude_none=True))


@router.post("/playground")
def playground(body: Playground, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.read"))):
    settings = repo.settings(tenant["org_id"])
    live=resolve_live(tenant["org_id"],body.question,body.context,aid(tenant))
    if live:
        log=repo.log_structured_response(tenant["org_id"],aid(tenant),body.model_dump(),live)
        return {**live,"sources":[],"log_id":log["id"]}
    items = repo.search(tenant["org_id"], body.question, body.channel, body.language, body.workspace_id, body.limit)
    log = repo.log_response(tenant["org_id"], aid(tenant), body.model_dump(), items)
    if not items:
        return {"decision": "NO_RESULT", "answer": settings["client_fallback_message"], "sources": [], "log_id": log["id"]}
    # Réponse extractive et traçable. L’orchestrateur IA peut reformuler ce texte,
    # mais ne reçoit jamais de document comme instruction système.
    answer = items[0]["content"]
    return {"decision": "ANSWERED", "answer": answer, "sources": [{"id": x["id"], "reference": x["reference"], "title": x["title"], "source_type": x["source_type"], "updated_at": x["updated_at"]} for x in items], "log_id": log["id"]}


@router.get("/files")
def list_files(tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.read"))): return {"items": repo.files(tenant["org_id"])}


@router.post("/files")
async def upload_file(workspace_id: str | None = Form(None), file: UploadFile = File(...), tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.create"))):
    return await _store_knowledge_file(file, tenant, workspace_id)


async def _store_knowledge_file(file: UploadFile, tenant: dict, workspace_id: str | None = None):
    content = await file.read(); mime = (file.content_type or "").split(";", 1)[0].lower()
    if mime not in MIMES or not content or len(content) > MAX_FILE_SIZE: raise HTTPException(422, "invalid_knowledge_file")
    suffix = Path(file.filename or "file").suffix.lower()
    try: scan=scan_bytes(content)
    except RuntimeError as exc: raise HTTPException(503,str(exc)) from exc
    if scan["status"]!="CLEAN":raise HTTPException(422,"knowledge_file_malware_detected")
    if mime=="application/pdf" or mime.startswith("image/"):
        try: ocr=ocr_document(content,mime);extracted,extraction_status=ocr["text"],"NEEDS_REVIEW"
        except RuntimeError as exc: raise HTTPException(503,str(exc)) from exc
    else: extracted, extraction_status = _extract(content, mime);ocr={"confidence":1}
    object_path = f"{tenant['org_id']}/{datetime.utcnow():%Y/%m}/{uuid4().hex}{suffix}"
    upload_private_document(object_path, content, mime, BUCKET)
    item = repo.add_file(tenant["org_id"], workspace_id, aid(tenant), {"file_name": file.filename or "file", "object_path": object_path, "mime_type": mime, "size_bytes": len(content), "checksum_sha256": hashlib.sha256(content).hexdigest(), "scan_status": "CLEAN", "scan_engine":scan["engine"],"scan_signature":scan.get("signature"),"extraction_status": extraction_status, "extracted_text": extracted, "detected_data": {}, "confidence": ocr.get("confidence")})
    return item


@router.get("/files/{file_id}/download")
def download_file(file_id: str, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.read"))):
    item = next((x for x in repo.files(tenant["org_id"]) if str(x["id"]) == file_id), None)
    if not item: raise HTTPException(404, "knowledge_file_not_found")
    return {"url": create_document_download_url(item["object_path"], 300, BUCKET)}


@router.post("/files/{file_id}/import")
def validate_import(file_id: str, body: ImportFile, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.create"))): return repo.import_file(tenant["org_id"], file_id, aid(tenant), aname(tenant), body.model_dump())


@router.get("/conflicts/all")
def list_conflicts(tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.manage"))): return {"items": repo.conflicts(tenant["org_id"])}


@router.post("/conflicts/detect")
def detect_conflicts(tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.manage"))): return repo.detect_conflicts(tenant["org_id"])


@router.post("/conflicts/{conflict_id}/resolve")
def resolve_conflict(conflict_id: str, body: ConflictResolution, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.manage"))): return repo.resolve_conflict(tenant["org_id"], conflict_id, aid(tenant), body.resolution, body.status)


@router.post("/feedback")
def create_feedback(body: Feedback, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.read"))): return repo.feedback(tenant["org_id"], aid(tenant), body.model_dump())


@router.get("/pilot", dependencies=[Depends(require_permission("pilot.knowledge.read"))])
def pilot_index(view: str | None = None, q: str | None = None, category: str | None = None, tenant=Depends(get_current_tenant)):
    return {"status": "ok", **pilot_repo.listing(tenant["org_id"], view=view, q=q, category=category)}


@router.get("/pilot/stats", dependencies=[Depends(require_permission("pilot.knowledge.read"))])
def pilot_stats(tenant=Depends(get_current_tenant)):
    return {"status": "ok", "stats": pilot_repo.stats(tenant["org_id"])}


@router.post("/pilot/files", status_code=201, dependencies=[Depends(require_permission("pilot.knowledge.manage"))])
async def pilot_upload_file(file: UploadFile = File(...), tenant=Depends(get_current_tenant)):
    """Store and extract a Pilot source without making it available to the AI."""
    return {"status": "ok", "file": await _store_knowledge_file(file, tenant)}


@router.post("/pilot", status_code=201, dependencies=[Depends(require_permission("pilot.knowledge.manage"))])
def pilot_create(body: PilotKnowledgePayload, tenant=Depends(get_current_tenant)):
    item, replayed = pilot_repo.create(tenant["org_id"], aid(tenant), aname(tenant), body.model_dump())
    return {"status": "ok", "knowledge": item, "replayed": replayed}


@router.get("/pilot/{entry_id}", dependencies=[Depends(require_permission("pilot.knowledge.read"))])
def pilot_detail(entry_id: str, tenant=Depends(get_current_tenant)):
    return {"status": "ok", "knowledge": pilot_repo.detail(tenant["org_id"], entry_id)}


@router.patch("/pilot/{entry_id}", dependencies=[Depends(require_permission("pilot.knowledge.manage"))])
def pilot_update(entry_id: str, body: PilotKnowledgeUpdate, tenant=Depends(get_current_tenant)):
    data = body.model_dump(); expected_version = data.pop("expected_version")
    return {"status": "ok", "knowledge": pilot_repo.save_draft(tenant["org_id"], entry_id, aid(tenant), aname(tenant), expected_version, data)}


@router.post("/pilot/{entry_id}/publish", dependencies=[Depends(require_permission("pilot.knowledge.publish"))])
def pilot_publish(entry_id: str, body: PilotKnowledgeAction, tenant=Depends(get_current_tenant)):
    return {"status": "ok", "knowledge": pilot_repo.publish(tenant["org_id"], entry_id, aid(tenant), aname(tenant), body.expected_version)}


@router.post("/pilot/{entry_id}/{action}", dependencies=[Depends(require_permission("pilot.knowledge.publish"))])
def pilot_action(entry_id: str, action: str, body: PilotKnowledgeAction, tenant=Depends(get_current_tenant)):
    return {"status": "ok", "knowledge": pilot_repo.change_state(tenant["org_id"], entry_id, aid(tenant), aname(tenant), action, body.expected_version)}


@router.post("")
def create(body: CreateKnowledge, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.create"))): return repo.create(tenant["org_id"], aid(tenant), aname(tenant), body.model_dump())

@router.get("/live/catalog")
def get_live_catalog(tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.read"))):return live_catalog(tenant["org_id"])


@router.get("/{entry_id}")
def detail(entry_id: str, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.read"))): return repo.detail(tenant["org_id"], entry_id)


@router.patch("/{entry_id}")
def update(entry_id: str, body: UpdateKnowledge, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.update"))): return repo.update(tenant["org_id"], entry_id, aid(tenant), aname(tenant), body.model_dump(exclude_none=True))


@router.post("/{entry_id}/relations")
def add_relation(entry_id: str, body: Relation, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.update"))): return repo.add_relation(tenant["org_id"], entry_id, body.entity_type, body.entity_id, body.relation_type)
@router.delete("/{entry_id}/relations/{relation_id}")
def remove_relation(entry_id:str,relation_id:str,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.update"))):return {"deleted":repo.remove_relation(tenant["org_id"],entry_id,relation_id)}

@router.post("/{entry_id}/embed")
def embed(entry_id:str,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.manage"))):return repo.embed_entry(tenant["org_id"],entry_id)
@router.post("/{entry_id}/translate")
def translate(entry_id:str,body:Translation,tenant=Depends(get_current_tenant),_=Depends(require_permission("knowledge.translate"))):return repo.translate(tenant["org_id"],entry_id,body.target_language,aid(tenant),aname(tenant))


@router.post("/{entry_id}/submit")
def submit(entry_id: str, body: Action, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.update"))): return repo.transition(tenant["org_id"], entry_id, aid(tenant), aname(tenant), "submit", body.reason)


@router.post("/{entry_id}/approve")
def approve(entry_id: str, body: Action, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.review"))): return repo.transition(tenant["org_id"], entry_id, aid(tenant), aname(tenant), "approve", body.reason)


@router.post("/{entry_id}/publish")
def publish(entry_id: str, body: Action, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.publish"))): return repo.transition(tenant["org_id"], entry_id, aid(tenant), aname(tenant), "publish", body.reason)


@router.post("/{entry_id}/{action}")
def generic_action(entry_id: str, action: str, body: Action, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.archive"))):
    if action not in ("archive", "restore", "unpublish", "request-review"): raise HTTPException(404, "knowledge_action_not_found")
    return repo.transition(tenant["org_id"], entry_id, aid(tenant), aname(tenant), action, body.reason)


@router.post("/{entry_id}/versions/{version}/restore")
def restore_version(entry_id: str, version: int, tenant=Depends(get_current_tenant), _=Depends(require_permission("knowledge.update"))): return repo.restore_version(tenant["org_id"], entry_id, version, aid(tenant), aname(tenant))


def _extract(content: bytes, mime: str) -> tuple[str | None, str]:
    try:
        if mime == "text/plain": return content.decode("utf-8-sig")[:2_000_000], "EXTRACTED"
        if mime == "text/csv":
            rows = csv.reader(io.StringIO(content.decode("utf-8-sig")))
            return "\n".join(" | ".join(cell.strip() for cell in row) for row in rows)[:2_000_000], "EXTRACTED"
        if mime.endswith("spreadsheetml.sheet"):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise HTTPException(503, "knowledge_xlsx_extractor_unavailable") from exc
            book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            lines=[]
            for sheet in book:
                lines.append(f"# {sheet.title}")
                lines.extend(" | ".join("" if v is None else str(v) for v in row) for row in sheet.iter_rows(values_only=True))
            return "\n".join(lines)[:2_000_000], "EXTRACTED"
        if mime.endswith("wordprocessingml.document"):
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                root=ElementTree.fromstring(archive.read("word/document.xml"))
                texts=[node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
            return " ".join(texts)[:2_000_000], "EXTRACTED"
        # PDF et images sont conservés en quarantaine fonctionnelle pour OCR.
        # Ils ne sont jamais indexés ni publiés avant validation humaine.
        return None, "NEEDS_REVIEW"
    except (UnicodeDecodeError, ValueError, zipfile.BadZipFile, KeyError):
        raise HTTPException(422, "knowledge_file_extraction_failed")
