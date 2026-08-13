import csv
import hashlib
import io
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, Field, model_validator
from starlette.responses import Response, StreamingResponse

from app.core.tenant_context import get_current_tenant
from app.core.permissions import require_permission
from app.core.config import settings
from app.services.dossier_document_storage import create_document_download_url, upload_private_document
from app.services.package_label_ocr import read_package_label
from app.packages.repository import (
    ANOMALY_SEVERITIES,
    INVENTORY_STATUSES,
    NOTIFICATION_CHANNELS,
    PACKAGE_CONDITIONS,
    PACKAGE_SOURCES,
    PACKAGE_STATUSES,
    PACKAGE_TYPES,
    PAYMENT_STATUSES,
    VALIDATION_STATUSES,
    add_package_media,
    create_package,
    create_package_anomaly,
    create_package_notification,
    export_packages,
    get_package,
    import_packages,
    list_packages,
    package_stats,
    package_timeline,
    resolve_package_anomaly,
    update_package,
    move_package, weigh_package, add_package_note, set_package_checklist,
    create_package_document, get_package_document, archive_package,
    add_private_package_media, get_package_media,
    package_analytics,
    transition_package,quality_control,expectations,create_expectation,price_package,compatible_departures,delivery_proof,detect_package_alerts,
    package_saved_views,save_package_view,delete_package_view,package_alerts,resolve_package_alert,bulk_package_operation,
)


router = APIRouter()


class PackagePayload(BaseModel):
    dossier_id: str
    package_reference: str | None = Field(default=None, max_length=120)
    tracking_id: str | None = Field(default=None, max_length=120)
    source: str = "manual"
    package_type: str = "carton"
    description: str | None = Field(default=None, max_length=260)
    category: str | None = Field(default=None, max_length=120)
    status: str = "CREATED"
    validation_status: str = "PENDING"
    payment_status: str = "UNKNOWN"
    payment_clearance_status: str | None = None
    package_condition: str = "UNKNOWN"
    inventory_status: str = "NOT_STORED"
    warehouse_name: str | None = Field(default=None, max_length=160)
    warehouse_zone: str | None = Field(default=None, max_length=80)
    warehouse_rack: str | None = Field(default=None, max_length=80)
    warehouse_location: str | None = Field(default=None, max_length=160)
    origin_country: str | None = Field(default=None, max_length=80)
    origin_city: str | None = Field(default=None, max_length=80)
    destination_country: str | None = Field(default=None, max_length=80)
    destination_city: str | None = Field(default=None, max_length=80)
    service_type: str | None = Field(default=None, max_length=80)
    shipping_mode: str | None = Field(default=None, max_length=80)
    shipment_reference: str | None = Field(default=None, max_length=120)
    public_tracking_enabled: bool = True
    eta_at: str | None = None
    received_at: str | None = None
    dispatched_at: str | None = None
    delivered_at: str | None = None
    weight_kg: float | None = None
    volumetric_weight_kg: float | None = None
    length_cm: float | None = None
    width_cm: float | None = None
    height_cm: float | None = None
    volume_cbm: float | None = None
    pieces_count: int = Field(default=1, ge=1)
    declared_value: float | None = None
    declared_currency: str | None = Field(default=None, max_length=12)
    is_fragile: bool = False
    notes: str | None = None
    fees_total: float | None = None
    fees_paid: float | None = None
    currency: str | None = Field(default=None, max_length=12)
    barcode: str | None = Field(default=None, max_length=160)
    qr_code_value: str | None = Field(default=None, max_length=220)
    last_scan_location: str | None = Field(default=None, max_length=180)
    priority: str = Field(default="NORMAL",pattern="^(LOW|NORMAL|HIGH|URGENT)$")
    assigned_to: str|None=Field(default=None,max_length=200)
    supplier_name: str|None=Field(default=None,max_length=180)
    supplier_tracking: str|None=Field(default=None,max_length=180)
    shipping_mark: str|None=Field(default=None,max_length=180)
    order_number: str|None=Field(default=None,max_length=180)
    external_reference: str|None=Field(default=None,max_length=180)
    subcategory: str|None=Field(default=None,max_length=120)
    goods_classification: str=Field(default="ORDINARY_GOODS",pattern="^(ORDINARY_GOODS|SENSITIVE_GOODS|DANGEROUS_GOODS|FRAGILE|ELECTRONICS|BATTERY|LIQUID|FOOD|PHARMACEUTICAL|HIGH_VALUE)$")
    declared_weight_kg: float|None=Field(default=None,ge=0)
    receiving_mode: str|None=Field(default=None,pattern="^(MANUAL|BARCODE|QR|OCR|IMPORT|API)$")
    route_id: str|None=None
    shipping_service_id: str|None=None
    expected_at: str|None=None

    @model_validator(mode="after")
    def validate_package(self):
        _assert_allowed(self.status, PACKAGE_STATUSES, "invalid_status")
        _assert_allowed(self.package_condition, PACKAGE_CONDITIONS, "invalid_package_condition")
        _assert_allowed(self.inventory_status, INVENTORY_STATUSES, "invalid_inventory_status")
        _assert_allowed(self.validation_status, VALIDATION_STATUSES, "invalid_validation_status")
        _assert_allowed(self.payment_status, PAYMENT_STATUSES, "invalid_payment_status")
        _assert_allowed(self.source, PACKAGE_SOURCES, "invalid_source")
        _assert_allowed(self.package_type, PACKAGE_TYPES, "invalid_package_type")
        return self


class PackagePatchPayload(BaseModel):
    tracking_id: str | None = Field(default=None, max_length=120)
    source: str | None = None
    package_type: str | None = None
    description: str | None = Field(default=None, max_length=260)
    category: str | None = Field(default=None, max_length=120)
    status: str | None = None
    validation_status: str | None = None
    payment_status: str | None = None
    payment_clearance_status: str | None = None
    package_condition: str | None = None
    inventory_status: str | None = None
    warehouse_name: str | None = Field(default=None, max_length=160)
    warehouse_zone: str | None = Field(default=None, max_length=80)
    warehouse_rack: str | None = Field(default=None, max_length=80)
    warehouse_location: str | None = Field(default=None, max_length=160)
    origin_country: str | None = Field(default=None, max_length=80)
    origin_city: str | None = Field(default=None, max_length=80)
    destination_country: str | None = Field(default=None, max_length=80)
    destination_city: str | None = Field(default=None, max_length=80)
    service_type: str | None = Field(default=None, max_length=80)
    shipping_mode: str | None = Field(default=None, max_length=80)
    shipment_reference: str | None = Field(default=None, max_length=120)
    public_tracking_enabled: bool | None = None
    eta_at: str | None = None
    received_at: str | None = None
    dispatched_at: str | None = None
    delivered_at: str | None = None
    weight_kg: float | None = None
    volumetric_weight_kg: float | None = None
    length_cm: float | None = None
    width_cm: float | None = None
    height_cm: float | None = None
    volume_cbm: float | None = None
    pieces_count: int | None = Field(default=None, ge=1)
    declared_value: float | None = None
    declared_currency: str | None = Field(default=None, max_length=12)
    is_fragile: bool | None = None
    notes: str | None = None
    fees_total: float | None = None
    fees_paid: float | None = None
    currency: str | None = Field(default=None, max_length=12)
    barcode: str | None = Field(default=None, max_length=160)
    qr_code_value: str | None = Field(default=None, max_length=220)
    last_scan_location: str | None = Field(default=None, max_length=180)
    priority: str|None=Field(default=None,pattern="^(LOW|NORMAL|HIGH|URGENT)$")
    assigned_to: str|None=Field(default=None,max_length=200)
    supplier_name: str|None=Field(default=None,max_length=180)
    supplier_tracking: str|None=Field(default=None,max_length=180)
    shipping_mark: str|None=Field(default=None,max_length=180)
    order_number: str|None=Field(default=None,max_length=180)
    external_reference: str|None=Field(default=None,max_length=180)
    subcategory: str|None=Field(default=None,max_length=120)
    goods_classification: str|None=Field(default=None,pattern="^(ORDINARY_GOODS|SENSITIVE_GOODS|DANGEROUS_GOODS|FRAGILE|ELECTRONICS|BATTERY|LIQUID|FOOD|PHARMACEUTICAL|HIGH_VALUE)$")
    declared_weight_kg: float|None=Field(default=None,ge=0)
    receiving_mode: str|None=Field(default=None,pattern="^(MANUAL|BARCODE|QR|OCR|IMPORT|API)$")
    route_id: str|None=None
    shipping_service_id: str|None=None
    expected_at: str|None=None

    @model_validator(mode="after")
    def validate_patch(self):
        for value, allowed, detail in [
            (self.status, PACKAGE_STATUSES, "invalid_status"),
            (self.package_condition, PACKAGE_CONDITIONS, "invalid_package_condition"),
            (self.inventory_status, INVENTORY_STATUSES, "invalid_inventory_status"),
            (self.validation_status, VALIDATION_STATUSES, "invalid_validation_status"),
            (self.payment_status, PAYMENT_STATUSES, "invalid_payment_status"),
            (self.source, PACKAGE_SOURCES, "invalid_source"),
            (self.package_type, PACKAGE_TYPES, "invalid_package_type"),
        ]:
            if value is not None:
                _assert_allowed(value, allowed, detail)
        return self


class PackageMediaPayload(BaseModel):
    media_url: str = Field(min_length=4, max_length=900)
    media_type: str = Field(default="IMAGE", max_length=40)
    caption: str | None = Field(default=None, max_length=220)


class PackageAnomalyPayload(BaseModel):
    anomaly_type: str = Field(default="OTHER", max_length=80)
    severity: str = "MEDIUM"
    title: str = Field(min_length=2, max_length=180)
    description: str | None = None

    @model_validator(mode="after")
    def validate_anomaly(self):
        _assert_allowed(self.severity, ANOMALY_SEVERITIES, "invalid_severity")
        return self


class PackageAnomalyResolvePayload(BaseModel):
    notes: str | None = Field(default=None, max_length=500)


class PackageNotificationPayload(BaseModel):
    channel: str = "whatsapp"
    notification_type: str = Field(default="PACKAGE_UPDATE", max_length=80)
    recipient: str | None = Field(default=None, max_length=160)
    message: str = Field(min_length=2, max_length=1200)

    @model_validator(mode="after")
    def validate_notification(self):
        _assert_allowed(self.channel, NOTIFICATION_CHANNELS, "invalid_channel")
        return self

class PackageMovePayload(BaseModel):
    to_warehouse: str = Field(min_length=1,max_length=160)
    to_zone: str | None = Field(default=None,max_length=80)
    to_aisle: str | None = Field(default=None,max_length=80)
    to_shelf: str | None = Field(default=None,max_length=80)
    to_position: str | None = Field(default=None,max_length=80)
    reason: str | None = Field(default=None,max_length=500)

class PackageWeightPayload(BaseModel):
    weight_kg: float = Field(ge=0)
    source: str = Field(default="MANUAL",max_length=40)
    device_reference: str | None = Field(default=None,max_length=120)
    notes: str | None = Field(default=None,max_length=500)

class PackageNotePayload(BaseModel):
    body: str = Field(min_length=1,max_length=4000)

class PackageChecklistPayload(BaseModel):
    status: str
    @model_validator(mode="after")
    def validate_status(self):
        _assert_allowed(self.status,{"PENDING","COMPLETED","NOT_APPLICABLE"},"invalid_checklist_status")
        return self
class PackageTransitionPayload(BaseModel):new_status:str;expected_version:int=Field(ge=1);reason:str|None=Field(default=None,max_length=500)
class PackageQualityPayload(BaseModel):packaging_intact:bool|None=None;label_readable:bool|None=None;product_compliant:bool|None=None;quantity_compliant:bool|None=None;no_damage:bool|None=None;no_moisture:bool|None=None;result:str=Field(pattern="^(COMPLIANT|REVIEW|NON_COMPLIANT)$");notes:str|None=Field(default=None,max_length=1000)
class PackageExpectationPayload(BaseModel):client_id:str;dossier_id:str|None=None;supplier_tracking:str=Field(min_length=3,max_length=160);shipping_mark:str|None=None;order_number:str|None=None;description:str|None=None;expected_warehouse_id:str|None=None;expected_at:str|None=None
class PackagePricingPayload(BaseModel):service_id:str
class PackageDeliveryPayload(BaseModel):recipient_name:str=Field(min_length=2,max_length=180);recipient_document:str|None=None;signature_path:str|None=None;photo_path:str|None=None;otp_verified:bool=False
class PackageSavedViewPayload(BaseModel):name:str=Field(min_length=2,max_length=80);filters:dict=Field(default_factory=dict)
class PackageAlertResolvePayload(BaseModel):resolution:str=Field(min_length=2,max_length=1000)
class PackageBulkPayload(BaseModel):idempotency_key:str=Field(min_length=8,max_length=120);operation_type:str;package_ids:list[str]=Field(min_length=1,max_length=500);payload:dict=Field(default_factory=dict)


def _user_id(tenant: dict) -> str:
    return str(tenant.get("user_id") or "")


def _assert_allowed(value: str, allowed: set[str], detail: str):
    if value not in allowed:
        raise ValueError(detail)


def _validate_query_value(value: str | None, allowed: set[str], detail: str):
    if value and value not in allowed:
        raise HTTPException(status_code=422, detail=detail)


@router.get("/packages", dependencies=[Depends(require_permission("packages.read"))])
def packages_index(
    q: str | None = Query(default=None, max_length=120),
    status_filter: str | None = Query(default=None, alias="status"),
    condition: str | None = None,
    inventory_status: str | None = None,
    payment_clearance_status: str | None = None,
    payment_status: str | None = None,
    validation_status: str | None = None,
    package_type: str | None = None,
    source: str | None = None,
    dossier_id: str | None = None,
    client_id: str | None = None,
    warehouse: str|None=None, country:str|None=None, city:str|None=None, shipment_id:str|None=None,
    batch_id:str|None=None, zone:str|None=None, responsible:str|None=None, category:str|None=None,
    priority:str|None=None, fragile:bool|None=None, received_from:str|None=None, received_to:str|None=None,
    min_declared_value:float|None=Query(default=None,ge=0), max_declared_value:float|None=Query(default=None,ge=0),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=30, ge=1, le=100),
    sort: str = "updated_desc",
    tenant=Depends(get_current_tenant),
):
    _validate_query_value(status_filter, PACKAGE_STATUSES, "invalid_status")
    _validate_query_value(condition, PACKAGE_CONDITIONS, "invalid_package_condition")
    _validate_query_value(inventory_status, INVENTORY_STATUSES, "invalid_inventory_status")
    _validate_query_value(payment_clearance_status, PAYMENT_STATUSES, "invalid_payment_status")
    _validate_query_value(payment_status, PAYMENT_STATUSES, "invalid_payment_status")
    _validate_query_value(validation_status, VALIDATION_STATUSES, "invalid_validation_status")
    _validate_query_value(package_type, PACKAGE_TYPES, "invalid_package_type")
    _validate_query_value(source, PACKAGE_SOURCES, "invalid_source")
    response = list_packages(
        tenant["org_id"],
        q=q,
        status=status_filter,
        condition=condition,
        inventory_status=inventory_status,
        payment_clearance_status=payment_clearance_status,
        payment_status=payment_status,
        validation_status=validation_status,
        package_type=package_type,
        source=source,
        dossier_id=dossier_id,
        client_id=client_id,
        warehouse=warehouse,country=country,city=city,shipment_id=shipment_id,batch_id=batch_id,zone=zone,
        responsible=responsible,category=category,priority=priority,fragile=fragile,received_from=received_from,
        received_to=received_to,min_declared_value=min_declared_value,max_declared_value=max_declared_value,
        page=page,
        page_size=page_size,
        sort=sort,
    )
    return {"status": "ok", "count": len(response["items"]), "packages": response["items"], **response}


@router.get("/packages/stats", dependencies=[Depends(require_permission("packages.read"))])
def packages_stats(tenant=Depends(get_current_tenant)):
    return {"status": "ok", "stats": package_stats(tenant["org_id"])}

@router.get("/packages/analytics", dependencies=[Depends(require_permission("packages.read"))])
def packages_analytics(tenant=Depends(get_current_tenant)):
    return {"status":"ok","analytics":package_analytics(tenant["org_id"])}

@router.get("/packages/expected",dependencies=[Depends(require_permission("packages.read"))])
def packages_expected(status:str|None=None,tenant=Depends(get_current_tenant)):return {"status":"ok","items":expectations(tenant["org_id"],status)}
@router.post("/packages/expected",dependencies=[Depends(require_permission("packages.create"))])
def packages_expected_create(body:PackageExpectationPayload,tenant=Depends(get_current_tenant)):return {"status":"ok","expectation":create_expectation(tenant["org_id"],_user_id(tenant),body.model_dump())}
@router.post("/packages/alerts/detect",dependencies=[Depends(require_permission("packages.anomalies"))])
def packages_alert_detect(tenant=Depends(get_current_tenant)):return {"status":"ok","created":detect_package_alerts(tenant["org_id"])}
@router.get("/packages/alerts",dependencies=[Depends(require_permission("packages.read"))])
def packages_alerts(status:str|None=None,tenant=Depends(get_current_tenant)):return {"status":"ok","items":package_alerts(tenant["org_id"],status)}
@router.patch("/packages/alerts/{alert_id}",dependencies=[Depends(require_permission("packages.anomalies"))])
def packages_alert_resolve(alert_id:str,body:PackageAlertResolvePayload,tenant=Depends(get_current_tenant)):return {"status":"ok","alert":resolve_package_alert(tenant["org_id"],alert_id,_user_id(tenant),body.resolution)}
@router.get("/packages/views",dependencies=[Depends(require_permission("packages.read"))])
def packages_views(tenant=Depends(get_current_tenant)):return {"status":"ok","items":package_saved_views(tenant["org_id"],_user_id(tenant))}
@router.post("/packages/views",dependencies=[Depends(require_permission("packages.read"))])
def packages_view_save(body:PackageSavedViewPayload,tenant=Depends(get_current_tenant)):return {"status":"ok","view":save_package_view(tenant["org_id"],_user_id(tenant),body.name,body.filters)}
@router.delete("/packages/views/{view_id}",dependencies=[Depends(require_permission("packages.read"))])
def packages_view_delete(view_id:str,tenant=Depends(get_current_tenant)):return {"status":"ok","deleted":delete_package_view(tenant["org_id"],_user_id(tenant),view_id)}
@router.post("/packages/bulk",dependencies=[Depends(require_permission("packages.bulk"))])
def packages_bulk(body:PackageBulkPayload,tenant=Depends(get_current_tenant)):return {"status":"ok","operation":bulk_package_operation(tenant["org_id"],_user_id(tenant),body.idempotency_key,body.operation_type,body.package_ids,body.payload)}

@router.get("/packages/archived", dependencies=[Depends(require_permission("packages.archive"))])
def packages_archived(q:str|None=Query(default=None,max_length=120),page:int=Query(default=1,ge=1),page_size:int=Query(default=30,ge=1,le=100),tenant=Depends(get_current_tenant)):
    response=list_packages(tenant["org_id"],q=q,page=page,page_size=page_size,archived=True)
    return {"status":"ok","count":len(response["items"]),"packages":response["items"],**response}


@router.get("/packages/export", dependencies=[Depends(require_permission("packages.export"))])
def packages_export(
    q: str | None = Query(default=None, max_length=120),
    status_filter: str | None = Query(default=None, alias="status"),
    condition: str | None = None,
    inventory_status: str | None = None,
    payment_clearance_status: str | None = None,
    payment_status: str | None = None,
    validation_status: str | None = None,
    package_type: str | None = None,
    source: str | None = None,
    sort: str = "updated_desc",
    tenant=Depends(get_current_tenant),
):
    rows = export_packages(
        tenant["org_id"],
        q=q,
        status=status_filter,
        condition=condition,
        inventory_status=inventory_status,
        payment_clearance_status=payment_clearance_status,
        payment_status=payment_status,
        validation_status=validation_status,
        package_type=package_type,
        source=source,
        sort=sort,
    )
    output = io.StringIO()
    fieldnames = [
        "package_reference",
        "tracking_id",
        "client_name",
        "dossier_reference",
        "source",
        "package_type",
        "description",
        "category",
        "status",
        "validation_status",
        "payment_status",
        "package_condition",
        "inventory_status",
        "warehouse_name",
        "warehouse_location",
        "origin_country",
        "origin_city",
        "destination_country",
        "destination_city",
        "service_type",
        "weight_kg",
        "volumetric_weight_kg",
        "length_cm",
        "width_cm",
        "height_cm",
        "volume_cbm",
        "pieces_count",
        "declared_value",
        "declared_currency",
        "fees_total",
        "fees_paid",
        "currency",
        "created_at",
        "updated_at",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in fieldnames})
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="slaivio-colis.csv"'},
    )


@router.post("/packages/import", dependencies=[Depends(require_permission("packages.import"))])
async def packages_import(file: UploadFile = File(...), tenant=Depends(get_current_tenant)):
    content = await file.read(10_485_761)
    if not content or len(content)>10_485_760: raise HTTPException(status_code=413,detail="import_file_too_large")
    suffix=Path(file.filename or "").suffix.lower()
    if suffix==".xlsx":
        try:
            from openpyxl import load_workbook
            workbook=load_workbook(io.BytesIO(content),read_only=True,data_only=True)
            sheet=workbook.active
            output=io.StringIO(); writer=csv.writer(output)
            for row in sheet.iter_rows(values_only=True): writer.writerow(["" if value is None else value for value in row])
            text=output.getvalue()
        except Exception as exc: raise HTTPException(status_code=422,detail="invalid_xlsx_file") from exc
    else:
        try: text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc: raise HTTPException(status_code=422, detail="invalid_csv_encoding") from exc
    return {"status": "ok", "result": import_packages(tenant["org_id"], _user_id(tenant), text)}


@router.post("/packages", status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_permission("packages.create"))])
def packages_create(body: PackagePayload, tenant=Depends(get_current_tenant)):
    try:
        package = create_package(tenant["org_id"], _user_id(tenant), body.model_dump())
    except ValueError as exc:
        if str(exc) == "dossier_required":
            raise HTTPException(status_code=422, detail="dossier_required") from exc
        if str(exc) == "dossier_not_found":
            raise HTTPException(status_code=404, detail="dossier_not_found") from exc
        raise
    return {"status": "ok", "package": package}


@router.get("/packages/{package_id}", dependencies=[Depends(require_permission("packages.read"))])
def packages_show(package_id: str, tenant=Depends(get_current_tenant)):
    package = get_package(tenant["org_id"], package_id)
    if not package:
        raise HTTPException(status_code=404, detail="package_not_found")
    return {"status": "ok", "package": package, "data": package}


@router.get("/packages/{package_id}/timeline", dependencies=[Depends(require_permission("packages.read"))])
def packages_timeline(package_id: str, tenant=Depends(get_current_tenant)):
    package = get_package(tenant["org_id"], package_id)
    if not package:
        raise HTTPException(status_code=404, detail="package_not_found")
    return {"status": "ok", "items": package_timeline(tenant["org_id"], package_id)}


@router.patch("/packages/{package_id}", dependencies=[Depends(require_permission("packages.update"))])
def packages_update(package_id: str, body: PackagePatchPayload, tenant=Depends(get_current_tenant)):
    package = update_package(
        tenant["org_id"],
        package_id,
        _user_id(tenant),
        body.model_dump(exclude_unset=True),
    )
    if not package:
        raise HTTPException(status_code=404, detail="package_not_found")
    return {"status": "ok", "package": package}


@router.post("/packages/{package_id}/media", dependencies=[Depends(require_permission("packages.update"))])
def packages_media_create(package_id: str, body: PackageMediaPayload, tenant=Depends(get_current_tenant)):
    package = add_package_media(tenant["org_id"], package_id, _user_id(tenant), body.model_dump())
    if not package:
        raise HTTPException(status_code=404, detail="package_not_found")
    return {"status": "ok", "package": package}


@router.post("/packages/{package_id}/anomalies", dependencies=[Depends(require_permission("packages.update"))])
def packages_anomaly_create(package_id: str, body: PackageAnomalyPayload, tenant=Depends(get_current_tenant)):
    package = create_package_anomaly(tenant["org_id"], package_id, _user_id(tenant), body.model_dump())
    if not package:
        raise HTTPException(status_code=404, detail="package_not_found")
    return {"status": "ok", "package": package}


@router.patch("/packages/{package_id}/anomalies/{anomaly_id}/resolve", dependencies=[Depends(require_permission("packages.update"))])
def packages_anomaly_resolve(package_id: str, anomaly_id: str, body: PackageAnomalyResolvePayload, tenant=Depends(get_current_tenant)):
    package = resolve_package_anomaly(tenant["org_id"], package_id, anomaly_id, _user_id(tenant), body.notes)
    if not package:
        raise HTTPException(status_code=404, detail="package_or_anomaly_not_found")
    return {"status": "ok", "package": package}


@router.post("/packages/{package_id}/notifications", dependencies=[Depends(require_permission("packages.update"))])
def packages_notification_create(package_id: str, body: PackageNotificationPayload, tenant=Depends(get_current_tenant)):
    package = create_package_notification(tenant["org_id"], package_id, _user_id(tenant), body.model_dump())
    if not package:
        raise HTTPException(status_code=404, detail="package_not_found")
    return {"status": "ok", "package": package}

@router.post("/packages/{package_id}/transition",dependencies=[Depends(require_permission("packages.update"))])
def packages_transition(package_id:str,body:PackageTransitionPayload,tenant=Depends(get_current_tenant)):return {"status":"ok","package":transition_package(tenant["org_id"],package_id,_user_id(tenant),body.new_status,body.expected_version,body.reason)}
@router.post("/packages/{package_id}/quality-control",dependencies=[Depends(require_permission("packages.quality"))])
def packages_quality(package_id:str,body:PackageQualityPayload,tenant=Depends(get_current_tenant)):return {"status":"ok","package":quality_control(tenant["org_id"],package_id,_user_id(tenant),body.model_dump())}
@router.post("/packages/{package_id}/pricing",dependencies=[Depends(require_permission("packages.pricing"))])
def packages_pricing(package_id:str,body:PackagePricingPayload,tenant=Depends(get_current_tenant)):return {"status":"ok","package":price_package(tenant["org_id"],package_id,_user_id(tenant),body.service_id)}
@router.get("/packages/{package_id}/compatible-departures",dependencies=[Depends(require_permission("packages.read"))])
def packages_compatible(package_id:str,tenant=Depends(get_current_tenant)):return {"status":"ok","items":compatible_departures(tenant["org_id"],package_id)}
@router.post("/packages/{package_id}/delivery-proof",dependencies=[Depends(require_permission("packages.delivery"))])
def packages_delivery(package_id:str,body:PackageDeliveryPayload,tenant=Depends(get_current_tenant)):return {"status":"ok","package":delivery_proof(tenant["org_id"],package_id,_user_id(tenant),body.model_dump())}

@router.post("/packages/{package_id}/move", dependencies=[Depends(require_permission("packages.update"))])
def packages_move(package_id: str, body: PackageMovePayload, tenant=Depends(get_current_tenant)):
    package=move_package(tenant["org_id"],package_id,_user_id(tenant),body.model_dump())
    if not package: raise HTTPException(status_code=404,detail="package_not_found")
    return {"status":"ok","package":package}

@router.post("/packages/{package_id}/weigh", dependencies=[Depends(require_permission("packages.update"))])
def packages_weigh(package_id: str, body: PackageWeightPayload, tenant=Depends(get_current_tenant)):
    package=weigh_package(tenant["org_id"],package_id,_user_id(tenant),body.model_dump())
    if not package: raise HTTPException(status_code=404,detail="package_not_found")
    return {"status":"ok","package":package}

@router.post("/packages/{package_id}/notes", dependencies=[Depends(require_permission("packages.update"))])
def packages_note(package_id: str, body: PackageNotePayload, tenant=Depends(get_current_tenant)):
    package=add_package_note(tenant["org_id"],package_id,_user_id(tenant),body.body)
    if not package: raise HTTPException(status_code=404,detail="package_not_found")
    return {"status":"ok","package":package}

@router.patch("/packages/{package_id}/checklist/{item_id}", dependencies=[Depends(require_permission("packages.update"))])
def packages_checklist(package_id: str,item_id: str,body: PackageChecklistPayload,tenant=Depends(get_current_tenant)):
    package=set_package_checklist(tenant["org_id"],package_id,item_id,_user_id(tenant),body.status)
    if not package: raise HTTPException(status_code=404,detail="package_or_checklist_not_found")
    return {"status":"ok","package":package}

@router.post("/packages/{package_id}/documents", dependencies=[Depends(require_permission("packages.update"))])
async def packages_document_upload(package_id: str,file: UploadFile=File(...),document_type: str=Form(default="OTHER"),notes: str|None=Form(default=None),tenant=Depends(get_current_tenant)):
    if not get_package(tenant["org_id"],package_id): raise HTTPException(status_code=404,detail="package_not_found")
    if file.content_type not in {"application/pdf","image/jpeg","image/png","image/webp"}: raise HTTPException(status_code=415,detail="unsupported_document_type")
    content=await file.read(settings.dossier_document_max_bytes+1)
    if not content or len(content)>settings.dossier_document_max_bytes: raise HTTPException(status_code=413,detail="document_too_large")
    safe_name=Path(file.filename or "document").name
    object_path=f"packages/{tenant['org_id']}/{package_id}/{uuid4().hex}-{safe_name}"
    try: upload_private_document(object_path,content,file.content_type,"package-documents")
    except RuntimeError as exc: raise HTTPException(status_code=503,detail=str(exc)) from exc
    document=create_package_document(tenant["org_id"],package_id,_user_id(tenant),{"document_type":document_type[:50].upper(),"file_name":safe_name,"object_path":object_path,"mime_type":file.content_type,"size_bytes":len(content),"checksum_sha256":hashlib.sha256(content).hexdigest(),"notes":notes[:500] if notes else None})
    return {"status":"ok","document":document}

@router.get("/packages/{package_id}/documents/{document_id}/download", dependencies=[Depends(require_permission("packages.read"))])
def packages_document_download(package_id: str,document_id: str,tenant=Depends(get_current_tenant)):
    document=get_package_document(tenant["org_id"],package_id,document_id)
    if not document: raise HTTPException(status_code=404,detail="document_not_found")
    try: url=create_document_download_url(document["object_path"],bucket_name="package-documents")
    except RuntimeError as exc: raise HTTPException(status_code=503,detail=str(exc)) from exc
    return {"status":"ok","url":url,"expires_in":300}

@router.delete("/packages/{package_id}", dependencies=[Depends(require_permission("packages.archive"))])
def packages_archive(package_id: str,tenant=Depends(get_current_tenant)):
    if not archive_package(tenant["org_id"],package_id,_user_id(tenant)): raise HTTPException(status_code=404,detail="package_not_found")
    return {"status":"ok"}

@router.post("/packages/{package_id}/restore", dependencies=[Depends(require_permission("packages.archive"))])
def packages_restore(package_id: str,tenant=Depends(get_current_tenant)):
    if not archive_package(tenant["org_id"],package_id,_user_id(tenant),True): raise HTTPException(status_code=404,detail="package_not_found")
    return {"status":"ok"}

@router.post("/packages/scan/ocr", dependencies=[Depends(require_permission("packages.create"))])
async def packages_scan_ocr(file:UploadFile=File(...),tenant=Depends(get_current_tenant)):
    if file.content_type not in {"image/jpeg","image/png","image/webp"}: raise HTTPException(status_code=415,detail="unsupported_ocr_image")
    content=await file.read(8_388_609)
    if not content or len(content)>8_388_608: raise HTTPException(status_code=413,detail="ocr_image_too_large")
    try: result=read_package_label(content,file.content_type)
    except RuntimeError as exc: raise HTTPException(status_code=503,detail=str(exc)) from exc
    except Exception as exc: raise HTTPException(status_code=502,detail="ocr_provider_failed") from exc
    return {"status":"ok","result":result,"requires_human_review":True}

@router.post("/packages/{package_id}/media/upload", dependencies=[Depends(require_permission("packages.update"))])
async def packages_media_upload(package_id:str,file:UploadFile=File(...),category:str=Form(default="RECEPTION"),caption:str|None=Form(default=None),tenant=Depends(get_current_tenant)):
    allowed={"image/jpeg","image/png","image/webp","video/mp4","audio/mpeg","audio/mp4","audio/ogg"}
    if file.content_type not in allowed: raise HTTPException(status_code=415,detail="unsupported_media_type")
    content=await file.read(26_214_401)
    if not content or len(content)>26_214_400: raise HTTPException(status_code=413,detail="media_too_large")
    safe_name=Path(file.filename or "media").name; object_path=f"{tenant['org_id']}/{package_id}/{uuid4().hex}-{safe_name}"
    try: upload_private_document(object_path,content,file.content_type,"package-media")
    except RuntimeError as exc: raise HTTPException(status_code=503,detail=str(exc)) from exc
    media_type="IMAGE" if file.content_type.startswith("image/") else "VIDEO" if file.content_type.startswith("video/") else "AUDIO"
    package=add_private_package_media(tenant["org_id"],package_id,_user_id(tenant),{"object_path":object_path,"file_name":safe_name,"mime_type":file.content_type,"size_bytes":len(content),"checksum_sha256":hashlib.sha256(content).hexdigest(),"category":category[:40].upper(),"caption":caption[:220] if caption else None,"media_type":media_type})
    if not package: raise HTTPException(status_code=404,detail="package_not_found")
    return {"status":"ok","package":package}

@router.get("/packages/{package_id}/media/{media_id}/view", dependencies=[Depends(require_permission("packages.read"))])
def packages_media_view(package_id:str,media_id:str,tenant=Depends(get_current_tenant)):
    media=get_package_media(tenant["org_id"],package_id,media_id)
    if not media: raise HTTPException(status_code=404,detail="media_not_found")
    try: url=create_document_download_url(media["object_path"],bucket_name="package-media")
    except RuntimeError as exc: raise HTTPException(status_code=503,detail=str(exc)) from exc
    return {"status":"ok","url":url,"expires_in":300}

@router.get("/packages/{package_id}/label/{kind}", dependencies=[Depends(require_permission("packages.read"))])
def packages_label(package_id:str,kind:str,tenant=Depends(get_current_tenant)):
    package=get_package(tenant["org_id"],package_id)
    if not package: raise HTTPException(status_code=404,detail="package_not_found")
    value=package.get("barcode") or package.get("tracking_id") or package.get("package_reference")
    if kind=="barcode":
        import barcode
        from barcode.writer import SVGWriter
        stream=io.BytesIO(); barcode.get("code128",value,writer=SVGWriter()).write(stream,{"write_text":True})
        return Response(stream.getvalue(),media_type="image/svg+xml",headers={"Content-Disposition":f'inline; filename="{package_id}-barcode.svg"'})
    if kind=="qr":
        import qrcode
        stream=io.BytesIO(); image=qrcode.make(package.get("qr_code_value") or value); image.save(stream,format="PNG")
        return Response(stream.getvalue(),media_type="image/png",headers={"Content-Disposition":f'inline; filename="{package_id}-qr.png"'})
    raise HTTPException(status_code=422,detail="invalid_label_kind")
